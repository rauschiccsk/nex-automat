"""
BaseGrid - Base class for all grids with advanced features.
PySide6 version for NEX Automat with enhanced functionality.
"""

from pathlib import Path
from typing import Any

from PySide6.QtCore import QModelIndex, Qt, Signal
from PySide6.QtGui import QAction, QBrush, QColor, QPainter
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QHeaderView,
    QInputDialog,
    QMenu,
    QMessageBox,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from shared_pyside6.database import SettingsRepository


class GreenHeaderView(QHeaderView):
    """
    Custom QHeaderView with green highlighting for active column.

    Features:
    - Highlights active search column with green background
    - Supports drag & drop column reordering
    """

    def __init__(self, orientation: Qt.Orientation, parent: QWidget | None = None):
        super().__init__(orientation, parent)
        self._active_column: int = -1
        self._green_color = QColor(144, 238, 144)  # Light green

        # Enable column moving
        self.setSectionsMovable(True)
        self.setSectionsClickable(True)

    def paintSection(self, painter: QPainter, rect, logicalIndex: int) -> None:
        """Custom paint for each header section."""
        painter.save()

        # Draw green background for active column
        if logicalIndex == self._active_column:
            painter.fillRect(rect, self._green_color)

        painter.restore()
        super().paintSection(painter, rect, logicalIndex)

    def set_active_column(self, column: int) -> None:
        """Set which column is active for search."""
        old_column = self._active_column
        self._active_column = column

        # Repaint affected sections
        if old_column >= 0:
            self.updateSection(old_column)
        if column >= 0:
            self.updateSection(column)


class BaseGrid(QWidget):
    """
    Base class for all grids in NEX Automat system.

    Features:
    - QTableView with GreenHeaderView
    - Column widths persistence
    - Column order persistence (drag & drop)
    - Column visibility (show/hide)
    - Custom headers (rename columns)
    - Active column persistence
    - Row cursor memory (remember selected row by ID)
    - Export to Excel/CSV
    - Multi-user support

    Usage:
        class MyGrid(BaseGrid):
            def __init__(self, parent=None):
                super().__init__(
                    window_name="my_window",
                    grid_name="my_grid",
                    parent=parent
                )
                self.model = MyModel()
                self.table_view.setModel(self.model)
                self.apply_model_and_load_settings()

    Signals:
        row_selected: Emitted when row selection changes
        row_activated: Emitted on double-click
        settings_changed: Emitted when grid settings change
    """

    # Signals
    row_selected = Signal(int)  # row index
    row_activated = Signal(int)  # row index
    settings_changed = Signal()

    def __init__(
        self,
        window_name: str,
        grid_name: str,
        user_id: str = "default",
        auto_load: bool = True,
        settings_db_path: str | Path | None = None,
        parent: QWidget | None = None,
    ):
        """
        Initialize BaseGrid.

        Args:
            window_name: Window identifier (e.g., "staging_main")
            grid_name: Grid identifier (e.g., "invoice_list")
            user_id: User ID for multi-user support
            auto_load: If True, automatically load settings after model set
            settings_db_path: Path to settings.db (default: ~/.nex-automat/settings.db)
            parent: Parent widget
        """
        super().__init__(parent)

        self._window_name = window_name
        self._grid_name = grid_name
        self._user_id = user_id
        self._auto_load = auto_load
        self._repository = SettingsRepository(settings_db_path)

        # Settings state
        self._column_widths: dict[str, int] = {}
        self._column_order: list[int] = []
        self._column_visibility: dict[str, bool] = {}
        self._custom_headers: dict[str, str] = {}
        self._active_column: int = 0
        self._last_row_id: Any = None
        self._row_id_column: int = 0
        self._sort_column: int | None = None
        self._sort_order: Qt.SortOrder = Qt.SortOrder.AscendingOrder

        # Debounce timer for saving
        self._save_pending = False

        # Flag to prevent saving during load (True until apply_model_and_load_settings)
        self._loading = True

        self._setup_base_ui()

    @property
    def window_name(self) -> str:
        return self._window_name

    @property
    def grid_name(self) -> str:
        return self._grid_name

    def _setup_base_ui(self) -> None:
        """Setup base UI with QTableView."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # Create table view with custom header
        self.table_view = QTableView()
        self.header = GreenHeaderView(Qt.Orientation.Horizontal, self.table_view)
        self.table_view.setHorizontalHeader(self.header)

        # Table view settings
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSortingEnabled(True)
        self.table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)

        # Vertical header (row numbers)
        self.table_view.verticalHeader().setVisible(False)

        layout.addWidget(self.table_view)

        # Enable context menu on header
        self.header.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.header.customContextMenuRequested.connect(self._show_header_context_menu)

        # Connect signals
        self._connect_base_signals()

    def _connect_base_signals(self) -> None:
        """Connect base signals."""
        # Header signals
        self.header.sectionResized.connect(self._on_column_resized)
        self.header.sectionMoved.connect(self._on_column_moved)
        self.header.sectionClicked.connect(self._on_header_clicked)

        # Selection signals
        self.table_view.clicked.connect(self._on_row_clicked)
        self.table_view.doubleClicked.connect(self._on_row_double_clicked)

        # Context menu
        self.table_view.customContextMenuRequested.connect(self._show_context_menu)

    # === Settings Load/Save ===

    def apply_model_and_load_settings(self) -> None:
        """
        Apply settings AFTER model is set.
        MUST be called by subclass AFTER self.table_view.setModel()!
        """
        if self._auto_load:
            self._load_grid_settings()
        self._loading = False  # Now allow saving

    def select_initial_row(self) -> None:
        """
        Select first row or restore last position.
        Call this AFTER populating model with data.
        """
        model = self.table_view.model()
        if not model or model.rowCount() == 0:
            return

        # Try to restore last position first
        if not self.restore_cursor_position():
            # Otherwise select first row
            self.table_view.selectRow(0)

    def _load_grid_settings(self) -> None:
        """Load and apply saved grid settings."""
        settings = self._repository.load_grid_settings(self._window_name, self._grid_name, self._user_id)

        if not settings:
            return

        model = self.table_view.model()
        if not model:
            return

        # Load column widths
        self._column_widths = settings.get("column_widths", {})
        for col_str, width in self._column_widths.items():
            col = int(col_str)
            if 0 <= col < model.columnCount():
                self.table_view.setColumnWidth(col, width)

        # Load column order
        self._column_order = settings.get("column_order", [])
        header = self.table_view.horizontalHeader()
        for visual_idx, logical_idx in enumerate(self._column_order):
            if logical_idx < model.columnCount():
                current_visual = header.visualIndex(logical_idx)
                if current_visual != visual_idx:
                    header.moveSection(current_visual, visual_idx)

        # Load column visibility
        self._column_visibility = settings.get("column_visibility", {})
        for col_str, visible in self._column_visibility.items():
            col = int(col_str)
            if 0 <= col < model.columnCount():
                self.table_view.setColumnHidden(col, not visible)

        # Load custom headers
        self._custom_headers = settings.get("custom_headers", {})

        # Apply custom headers to model
        for col_str, header_text in self._custom_headers.items():
            col = int(col_str)
            if 0 <= col < model.columnCount():
                model.setHeaderData(col, Qt.Orientation.Horizontal, header_text)

        # Load active column
        self._active_column = settings.get("active_column", 0)
        self.header.set_active_column(self._active_column)

        # Load last row ID
        self._last_row_id = settings.get("last_row_id")

        # Load sort settings
        self._sort_column = settings.get("sort_column")
        sort_order_str = settings.get("sort_order", "ascending")
        self._sort_order = (
            Qt.SortOrder.DescendingOrder if sort_order_str == "descending" else Qt.SortOrder.AscendingOrder
        )

        if self._sort_column is not None:
            self.table_view.sortByColumn(self._sort_column, self._sort_order)

    def _save_grid_settings(self) -> None:
        """Save current grid settings."""
        if self._loading:
            return  # Don't save during initialization

        model = self.table_view.model()
        if not model:
            return

        # Collect column widths
        column_widths = {}
        for col in range(model.columnCount()):
            column_widths[str(col)] = self.table_view.columnWidth(col)

        # Collect column order
        header = self.table_view.horizontalHeader()
        column_order = []
        for visual_idx in range(model.columnCount()):
            column_order.append(header.logicalIndex(visual_idx))

        # Collect visibility
        column_visibility = {}
        for col in range(model.columnCount()):
            column_visibility[str(col)] = not self.table_view.isColumnHidden(col)

        # Save last row ID
        self._save_cursor_position()

        settings = {
            "column_widths": column_widths,
            "column_order": column_order,
            "column_visibility": column_visibility,
            "custom_headers": self._custom_headers,
            "active_column": self._active_column,
            "last_row_id": self._last_row_id,
            "sort_column": self._sort_column,
            "sort_order": ("descending" if self._sort_order == Qt.SortOrder.DescendingOrder else "ascending"),
        }

        self._repository.save_grid_settings(self._window_name, self._grid_name, self._user_id, settings)

        self.settings_changed.emit()

    def save_grid_settings_now(self) -> None:
        """Manual save grid settings."""
        self._save_grid_settings()

    def reload_grid_settings(self) -> None:
        """Manual reload grid settings."""
        self._load_grid_settings()

    # === Event Handlers ===

    def _on_column_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        """Handle column resize."""
        self._column_widths[str(logical_index)] = new_size
        self._save_grid_settings()

    def _on_column_moved(self, logical_index: int, old_visual_index: int, new_visual_index: int) -> None:
        """Handle column move (drag & drop)."""
        self._save_grid_settings()

    def _on_header_clicked(self, logical_index: int) -> None:
        """Handle header click - set active column."""
        self._active_column = logical_index
        self.header.set_active_column(logical_index)
        self._save_grid_settings()

    def _on_row_clicked(self, index: QModelIndex) -> None:
        """Handle row click."""
        self.row_selected.emit(index.row())

    def _on_row_double_clicked(self, index: QModelIndex) -> None:
        """Handle row double-click."""
        self.row_activated.emit(index.row())

    # === Column Visibility ===

    def set_column_visible(self, column: int, visible: bool) -> None:
        """Set column visibility."""
        self.table_view.setColumnHidden(column, not visible)
        self._column_visibility[str(column)] = visible

        # Restore minimum width when making visible
        if visible:
            current_width = self.table_view.columnWidth(column)
            if current_width < 20:
                # Restore from saved width or use default
                saved_width = self._column_widths.get(str(column), 80)
                self.table_view.setColumnWidth(column, max(saved_width, 50))

        self._save_grid_settings()

    def is_column_visible(self, column: int) -> bool:
        """Check if column is visible."""
        return not self.table_view.isColumnHidden(column)

    def get_visible_columns(self) -> list[int]:
        """Get list of visible column indices."""
        model = self.table_view.model()
        if not model:
            return []
        return [col for col in range(model.columnCount()) if not self.table_view.isColumnHidden(col)]

    # === Custom Headers ===

    def set_custom_header(self, column: int, text: str) -> None:
        """Set custom header text for column."""
        self._custom_headers[str(column)] = text
        self._save_grid_settings()
        # Note: Model should check _custom_headers in headerData()

    def get_custom_header(self, column: int) -> str | None:
        """Get custom header text for column."""
        return self._custom_headers.get(str(column))

    def reset_headers(self) -> None:
        """Reset all custom headers to original."""
        self._custom_headers.clear()
        self._save_grid_settings()

    # === Row Cursor Memory ===

    def set_row_id_column(self, column: int) -> None:
        """Set which column contains unique row ID."""
        self._row_id_column = column

    def _save_cursor_position(self) -> None:
        """Save current cursor position (row ID)."""
        model = self.table_view.model()
        if not model:
            return

        current = self.table_view.currentIndex()
        if current.isValid():
            id_index = model.index(current.row(), self._row_id_column)
            self._last_row_id = model.data(id_index, Qt.ItemDataRole.DisplayRole)

    def restore_cursor_position(self) -> bool:
        """
        Restore cursor to last selected row by ID.

        Returns:
            True if restored successfully
        """
        if self._last_row_id is None:
            return False

        model = self.table_view.model()
        if not model:
            return False

        # Find row with matching ID
        for row in range(model.rowCount()):
            index = model.index(row, self._row_id_column)
            row_id = model.data(index, Qt.ItemDataRole.DisplayRole)
            if row_id == self._last_row_id:
                self.table_view.selectRow(row)
                self.table_view.scrollTo(index)
                return True

        return False

    # === Active Column ===

    def get_active_column(self) -> int:
        """Get active column index."""
        return self._active_column

    def set_active_column(self, column: int) -> None:
        """Set active column."""
        self._active_column = column
        self.header.set_active_column(column)
        self._save_grid_settings()

    # === Item Creation ===

    def create_item(self, value, editable: bool = False) -> "QStandardItem":
        """
        Create QStandardItem with automatic formatting and alignment.

        - Integers: right-aligned, no decimal places
        - Floats: right-aligned, 2 decimal places
        - Strings/other: left-aligned

        Args:
            value: The value to display
            editable: Whether the item is editable

        Returns:
            Configured QStandardItem
        """
        from PySide6.QtGui import QStandardItem

        # Determine text and alignment based on type
        is_boolean = False
        if value is None:
            text = ""
            align_right = False
        elif isinstance(value, bool):
            # Use icons for boolean: green checkmark / red X
            text = "✓" if value else "✗"
            is_boolean = True
            align_right = False
        elif isinstance(value, float):
            # Float always with 2 decimal places (including 0.0)
            text = f"{value:.2f}"
            align_right = True
        elif isinstance(value, int):
            text = str(value)
            align_right = True
        else:
            # Try to detect numeric strings
            text = str(value)
            align_right = False

        item = QStandardItem(text)
        item.setEditable(editable)

        if align_right:
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        else:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter) if is_boolean else None

        # Set color for boolean icons
        if is_boolean:
            if value:
                item.setForeground(QBrush(QColor(0, 200, 0)))  # Green
            else:
                item.setForeground(QBrush(QColor(220, 50, 50)))  # Red

        return item

    # === Export ===

    def export_to_csv(self, filepath: str | None = None) -> bool:
        """
        Export visible data to CSV.

        Args:
            filepath: Path to save file. If None, shows file dialog.

        Returns:
            True if successful
        """
        if filepath is None:
            filepath, _ = QFileDialog.getSaveFileName(self, "Export to CSV", "", "CSV Files (*.csv)")
            if not filepath:
                return False

        model = self.table_view.model()
        if not model:
            return False

        try:
            visible_cols = self.get_visible_columns()

            with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
                import csv

                writer = csv.writer(f, delimiter=";")

                # Headers
                headers = []
                for col in visible_cols:
                    custom = self.get_custom_header(col)
                    if custom:
                        headers.append(custom)
                    else:
                        headers.append(model.headerData(col, Qt.Orientation.Horizontal) or "")
                writer.writerow(headers)

                # Data
                for row in range(model.rowCount()):
                    row_data = []
                    for col in visible_cols:
                        index = model.index(row, col)
                        value = model.data(index, Qt.ItemDataRole.DisplayRole)
                        row_data.append(str(value) if value is not None else "")
                    writer.writerow(row_data)

            return True
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {e}")
            return False

    def export_to_excel(self, filepath: str | None = None) -> bool:
        """
        Export visible data to Excel.

        Args:
            filepath: Path to save file. If None, shows file dialog.

        Returns:
            True if successful
        """
        if filepath is None:
            filepath, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "", "Excel Files (*.xlsx)")
            if not filepath:
                return False

        model = self.table_view.model()
        if not model:
            return False

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font

            visible_cols = self.get_visible_columns()

            wb = Workbook()
            ws = wb.active
            ws.title = self._grid_name

            # Headers (bold)
            for col_idx, col in enumerate(visible_cols, 1):
                custom = self.get_custom_header(col)
                if custom:
                    header_text = custom
                else:
                    header_text = model.headerData(col, Qt.Orientation.Horizontal) or ""
                cell = ws.cell(row=1, column=col_idx, value=header_text)
                cell.font = Font(bold=True)

            # Data
            for row in range(model.rowCount()):
                for col_idx, col in enumerate(visible_cols, 1):
                    index = model.index(row, col)
                    value = model.data(index, Qt.ItemDataRole.DisplayRole)
                    ws.cell(row=row + 2, column=col_idx, value=value)

            wb.save(filepath)
            return True
        except ImportError:
            QMessageBox.critical(self, "Export Error", "openpyxl not installed. Run: pip install openpyxl")
            return False
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {e}")
            return False

    # === Context Menu ===

    def _show_header_context_menu(self, position) -> None:
        """Show context menu on header for column visibility and rename."""
        model = self.table_view.model()
        if not model:
            return

        # Get column under cursor
        clicked_col = self.header.logicalIndexAt(position)

        menu = QMenu(self)

        # Rename option for clicked column
        if clicked_col >= 0:
            current_name = self.get_custom_header(clicked_col)
            if not current_name:
                current_name = model.headerData(clicked_col, Qt.Orientation.Horizontal) or f"Column {clicked_col}"

            rename_action = QAction(f"Premenovať '{current_name}'...", self)
            rename_action.triggered.connect(lambda: self._rename_column_dialog(clicked_col))
            menu.addAction(rename_action)

            # Reset name option (only if custom name exists)
            if self.get_custom_header(clicked_col):
                reset_action = QAction("Obnoviť pôvodný názov", self)
                reset_action.triggered.connect(lambda: self._reset_column_name(clicked_col))
                menu.addAction(reset_action)

            menu.addSeparator()

        # Column visibility submenu
        columns_menu = menu.addMenu("Stĺpce")
        for col in range(model.columnCount()):
            # Get display name (custom or original)
            custom = self.get_custom_header(col)
            if custom:
                header = custom
            else:
                header = model.headerData(col, Qt.Orientation.Horizontal) or f"Column {col}"

            action = QAction(str(header), self)
            action.setCheckable(True)
            action.setChecked(self.is_column_visible(col))
            action.triggered.connect(lambda checked, c=col: self.set_column_visible(c, checked))
            columns_menu.addAction(action)

        menu.exec(self.header.mapToGlobal(position))

    def _rename_column_dialog(self, column: int) -> None:
        """Show dialog to rename column."""
        model = self.table_view.model()
        if not model:
            return

        # Get current name
        current = self.get_custom_header(column)
        if not current:
            current = model.headerData(column, Qt.Orientation.Horizontal) or ""

        # Show input dialog
        new_name, ok = QInputDialog.getText(self, "Premenovať stĺpec", "Nový názov:", text=str(current))

        if ok and new_name:
            self.set_custom_header(column, new_name)
            # Update header display
            model.setHeaderData(column, Qt.Orientation.Horizontal, new_name)

    def _reset_column_name(self, column: int) -> None:
        """Reset column name to original."""
        model = self.table_view.model()
        if not model or str(column) not in self._custom_headers:
            return

        # Remove custom header
        del self._custom_headers[str(column)]
        self._save_grid_settings()

        # Restore original - need to get from model's original data
        # For QStandardItemModel, we need to reload or store originals
        # For now, just trigger a visual update
        self.header.updateSection(column)

    def _show_context_menu(self, position) -> None:
        """Show context menu."""
        menu = QMenu(self)

        # Export actions
        export_csv = QAction("Export to CSV...", self)
        export_csv.triggered.connect(lambda: self.export_to_csv())
        menu.addAction(export_csv)

        export_excel = QAction("Export to Excel...", self)
        export_excel.triggered.connect(lambda: self.export_to_excel())
        menu.addAction(export_excel)

        menu.addSeparator()

        # Column visibility submenu
        columns_menu = menu.addMenu("Columns")
        model = self.table_view.model()
        if model:
            for col in range(model.columnCount()):
                header = model.headerData(col, Qt.Orientation.Horizontal) or f"Column {col}"
                action = QAction(str(header), self)
                action.setCheckable(True)
                action.setChecked(self.is_column_visible(col))
                action.triggered.connect(lambda checked, c=col: self.set_column_visible(c, checked))
                columns_menu.addAction(action)

        menu.exec(self.table_view.viewport().mapToGlobal(position))
